#!/usr/bin/env python3
"""
polymarket_price_updater.py — Active Positions Price Sync
Pierre & Gabe | v1.0 | April 2026

Fetches current market prices for all Open positions in your tracker's
Active Positions tab and updates the "Current Price ($)" column.

Also flags any positions where the current price has breached their stop-loss.

Usage:
  python polymarket_price_updater.py
  python polymarket_price_updater.py --tracker path/to/tracker.xlsx
  python polymarket_price_updater.py --dry-run   # show what would change, don't save

How market matching works:
  The script looks for the Polymarket market ID or URL in the Notes column
  of Active Positions (placed there automatically if you used the scanner to
  find the market). If no ID is found in Notes, it falls back to fuzzy text
  matching on the Market Question.

  Best practice: When you log a position, paste the Polymarket URL in the
  Notes column. Format: https://polymarket.com/event/some-slug
  The script extracts the slug and uses it to find the right market.

Requirements:
  pip install requests openpyxl
"""

import requests
import json
import argparse
import os
import re
import sys
import time
from datetime import datetime, timezone
from difflib import SequenceMatcher

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Error: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

GAMMA_API = "https://gamma-api.polymarket.com"

# ─── COLUMN MAPPING: Active Positions tab (1-based, openpyxl) ─────────────────
# Adjust these if you ever restructure the spreadsheet.
AP_COLS = {
    "id":             1,   # A — Position ID
    "question":       2,   # B — Market Question
    "direction":      3,   # C — YES/NO
    "entry_date":     4,   # D — Entry Date
    "resolution":     5,   # E — Expected Resolution
    "entry_price":    6,   # F — Entry Price ($)
    "current_price":  7,   # G — Current Price ($)
    "shares":         8,   # H — Shares
    "deployed":       9,   # I — Capital Deployed ($)
    "our_prob":       10,  # J — Our Prob (%)
    "market_prob":    11,  # K — Market Prob at Entry (%)
    "edge":           12,  # L — Edge (pp) — auto-calculated in sheet
    "confidence":     13,  # M — Confidence
    "category":       14,  # N — Category
    "thesis":         15,  # O — Thesis
    "assumptions":    16,  # P — Key Assumptions
    "kill":           17,  # Q — Kill Conditions
    "stop_loss":      18,  # R — Stop-Loss Price ($)
    "correlated":     19,  # S — Correlated Positions
    "status":         20,  # T — Status
    "notes":          21,  # U — Notes
}


# ─── API ───────────────────────────────────────────────────────────────────────

def search_market_by_id(market_id: str) -> dict | None:
    """Fetch a single market by its Polymarket ID."""
    try:
        resp = requests.get(
            f"{GAMMA_API}/markets/{market_id}",
            timeout=15,
        )
        if resp.status_code == 200:
            return resp.json()
    except requests.exceptions.RequestException:
        pass
    return None


def search_market_by_slug(slug: str) -> dict | None:
    """Fetch markets for an event by its slug."""
    try:
        resp = requests.get(
            f"{GAMMA_API}/events",
            params={"slug": slug, "limit": 5},
            timeout=15,
        )
        if resp.status_code == 200:
            events = resp.json()
            for event in events:
                markets = event.get("markets", [])
                if markets:
                    return markets[0]  # Return first market in the event
    except requests.exceptions.RequestException:
        pass
    return None


def search_market_by_question(question: str, top_n: int = 20) -> dict | None:
    """
    Fall-back: search by question text similarity.
    Fetches the top N markets by volume and returns the best fuzzy match.
    """
    try:
        resp = requests.get(
            f"{GAMMA_API}/markets",
            params={
                "active": "true",
                "limit": top_n,
                "order": "volume",
                "ascending": "false",
            },
            timeout=20,
        )
        if resp.status_code != 200:
            return None

        markets = resp.json()
        best_match = None
        best_score = 0.0

        q_lower = question.lower()
        for m in markets:
            mq = m.get("question", "").lower()
            score = SequenceMatcher(None, q_lower, mq).ratio()
            if score > best_score:
                best_score = score
                best_match = m

        # Only accept if similarity is reasonably high
        if best_score >= 0.70:
            return best_match

    except requests.exceptions.RequestException:
        pass
    return None


def get_current_yes_price(market: dict) -> float | None:
    """Extract YES price from a market dict."""
    try:
        raw = market.get("outcomePrices", "[]")
        prices = json.loads(raw) if isinstance(raw, str) else raw
        if prices:
            return float(prices[0])
    except (json.JSONDecodeError, ValueError, TypeError, IndexError):
        pass
    return None


# ─── NOTES PARSING ─────────────────────────────────────────────────────────────

def extract_id_from_notes(notes: str) -> tuple[str | None, str | None]:
    """
    Try to extract a Polymarket market ID or slug from the Notes field.
    Returns (market_id, slug) — either may be None.

    Handles formats like:
      https://polymarket.com/event/some-slug
      https://polymarket.com/market/0x1234abcd...
      ID:0x1234abcd...
    """
    if not notes:
        return None, None

    # Look for event slug URL
    slug_match = re.search(r'polymarket\.com/event/([a-z0-9\-]+)', notes, re.IGNORECASE)
    if slug_match:
        return None, slug_match.group(1)

    # Look for market URL with hex ID
    id_match = re.search(r'polymarket\.com/market/(0x[a-fA-F0-9]+)', notes, re.IGNORECASE)
    if id_match:
        return id_match.group(1), None

    # Look for explicit ID: tag
    explicit_match = re.search(r'\bID:(0x[a-fA-F0-9]+|\d+)', notes, re.IGNORECASE)
    if explicit_match:
        return explicit_match.group(1), None

    return None, None


# ─── STOP-LOSS CHECK ───────────────────────────────────────────────────────────

def check_stop_loss(
    question: str,
    direction: str,
    current_price: float,
    stop_loss: float,
) -> bool:
    """
    Returns True if the current price has breached the stop-loss.
    For YES bets: breach if current_price <= stop_loss
    For NO bets: breach if current_price >= stop_loss (price risen against us)
    """
    if direction == "YES":
        return current_price <= stop_loss
    elif direction == "NO":
        # On a NO bet, we lose if YES price rises (NO price falls)
        # stop_loss is stored as the YES price at which we exit
        return current_price >= stop_loss
    return False


# ─── MAIN LOGIC ────────────────────────────────────────────────────────────────

def process_positions(tracker_path: str, dry_run: bool = False) -> None:
    """Read Active Positions, fetch prices, update spreadsheet, flag stop-losses."""

    wb = load_workbook(tracker_path)

    if "Active Positions" not in wb.sheetnames:
        print("'Active Positions' sheet not found.")
        return

    ws = wb["Active Positions"]

    # Collect Open positions (skip header row 1)
    positions = []
    for row_idx in range(2, ws.max_row + 1):
        row_id = ws.cell(row_idx, AP_COLS["id"]).value
        question = ws.cell(row_idx, AP_COLS["question"]).value
        status = ws.cell(row_idx, AP_COLS["status"]).value
        direction = ws.cell(row_idx, AP_COLS["direction"]).value
        stop_loss = ws.cell(row_idx, AP_COLS["stop_loss"]).value
        notes = ws.cell(row_idx, AP_COLS["notes"]).value

        # Skip empty rows and non-Open positions
        if not question:
            continue
        if str(status).strip().lower() not in ("open", ""):
            continue

        positions.append({
            "row": row_idx,
            "id": row_id,
            "question": str(question).strip(),
            "direction": str(direction).strip().upper() if direction else "YES",
            "stop_loss": float(stop_loss) if stop_loss else None,
            "notes": str(notes) if notes else "",
        })

    if not positions:
        print("No open positions found in Active Positions tab.")
        return

    print(f"Found {len(positions)} open position(s). Fetching current prices...\n")

    # Fonts for highlighting
    alert_font = Font(name="Arial", size=10, bold=True, color="C0392B")
    normal_font = Font(name="Arial", size=10)
    alert_fill = PatternFill("solid", start_color="FDEDEC")

    stop_loss_alerts = []
    updated_count = 0

    for pos in positions:
        print(f"  [{pos['id']}] {pos['question'][:60]}")

        # Try to find the market
        market = None
        match_method = "none"

        market_id, slug = extract_id_from_notes(pos["notes"])

        if market_id:
            market = search_market_by_id(market_id)
            match_method = "ID"
        if not market and slug:
            market = search_market_by_slug(slug)
            match_method = "slug"
        if not market:
            market = search_market_by_question(pos["question"])
            match_method = "fuzzy text"

        if not market:
            print(f"    ⚠ Could not find on Polymarket — skipping")
            print(f"      Tip: Add the Polymarket URL to the Notes column for this position")
            continue

        current_price = get_current_yes_price(market)
        if current_price is None:
            print(f"    ⚠ Price not available from API")
            continue

        print(f"    Current price: {current_price:.3f} ({current_price*100:.1f}%) "
              f"[matched via {match_method}]")

        # Stop-loss check
        sl_breached = False
        if pos["stop_loss"] is not None:
            sl_breached = check_stop_loss(
                pos["question"],
                pos["direction"],
                current_price,
                pos["stop_loss"],
            )
            if sl_breached:
                print(f"    🚨 STOP-LOSS BREACHED! "
                      f"Current: {current_price:.3f} | Stop: {pos['stop_loss']:.3f}")
                stop_loss_alerts.append(pos)

        if not dry_run:
            cell = ws.cell(pos["row"], AP_COLS["current_price"])
            cell.value = current_price
            cell.number_format = "$#,##0.00"

            if sl_breached:
                cell.font = alert_font
                cell.fill = alert_fill
                # Also highlight the stop-loss cell
                ws.cell(pos["row"], AP_COLS["stop_loss"]).font = alert_font
                ws.cell(pos["row"], AP_COLS["stop_loss"]).fill = alert_fill
            else:
                cell.font = normal_font

        updated_count += 1
        time.sleep(0.3)  # Rate limiting

    # Save
    if not dry_run and updated_count > 0:
        wb.save(tracker_path)
        print(f"\n✓ Saved. Updated {updated_count} position(s) in {tracker_path}")
    elif dry_run:
        print(f"\n[DRY RUN] Would have updated {updated_count} position(s).")

    # Print stop-loss summary
    if stop_loss_alerts:
        print(f"\n{'='*55}")
        print(f"  🚨 STOP-LOSS ALERT — {len(stop_loss_alerts)} position(s)")
        print(f"{'='*55}")
        for pos in stop_loss_alerts:
            print(f"\n  Position [{pos['id']}]: {pos['question']}")
            print(f"  Direction: {pos['direction']}")
            print(f"  Stop-loss: ${pos['stop_loss']:.3f}")
            print(f"  Action required: Review at next session.")
            print(f"  Per your rules: stop-loss hit → exit immediately.")
        print(f"\n  Text Pierre if you need to exit before next session.")
        print(f"{'='*55}\n")
    else:
        print("\n  ✓ No stop-losses breached.")


def main():
    parser = argparse.ArgumentParser(
        description="Update current prices in Active Positions from Polymarket API"
    )
    parser.add_argument(
        "--tracker", default="polymarket_tracker_v2.xlsx",
        help="Path to tracker xlsx (default: polymarket_tracker_v2.xlsx)"
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Show what would change without saving to the spreadsheet"
    )
    args = parser.parse_args()

    run_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    print(f"\n{'='*55}")
    print(f"  Price Updater | {run_date}")
    if args.dry_run:
        print("  [DRY RUN — no changes will be saved]")
    print(f"{'='*55}\n")

    if not os.path.exists(args.tracker):
        print(f"Tracker not found: '{args.tracker}'")
        print("Use --tracker path/to/your/tracker.xlsx")
        sys.exit(1)

    process_positions(args.tracker, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
