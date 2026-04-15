#!/usr/bin/env python3
"""
polymarket_scanner.py — Polymarket Geopolitical Bet Discovery Pipeline
Pierre & Gabe | v1.0 | April 2026

Scans Polymarket for markets matching your geopolitical edge, produces:
  1. An HTML digest report for manual review
  2. Pipeline candidates written into the Bet Pipeline tab of your tracker xlsx

Usage:
  python polymarket_scanner.py
  python polymarket_scanner.py --tracker path/to/tracker.xlsx --days 120 --top 50
  python polymarket_scanner.py --no-spreadsheet   # report only, don't touch the xlsx
  python polymarket_scanner.py --days 180         # extended scan for longer-dated markets

Arguments:
  --tracker PATH         Path to your tracker xlsx (default: polymarket_tracker_v2.xlsx)
  --output-dir DIR       Where to save the HTML report (default: current directory)
  --days N               Max days to resolution to include (default: 90)
  --min-liquidity N      Min USD liquidity (default: 500)
  --min-volume N         Min USD volume traded (default: 1000)
  --top N                Max candidates to include in output (default: 40)
  --no-spreadsheet       Skip writing to spreadsheet; produce report only

Requirements:
  pip install requests openpyxl
"""

import requests
import json
import argparse
import os
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Warning: openpyxl not installed. Spreadsheet update will be skipped.")
    print("Install with: pip install openpyxl\n")


# ─── CONFIGURATION ─────────────────────────────────────────────────────────────

GAMMA_API = "https://gamma-api.polymarket.com"

# Keyword groups aligned with your edge areas.
# Edit freely — these drive which markets surface in the scan.
KEYWORD_GROUPS = {
    "Iran": [
        "iran", "iaea", "nuclear deal", "jcpoa", "kharg", "hormuz",
        "ayatollah", "khamenei", "tehran", "iranian", "persian gulf",
        "iran nuclear", "us iran", "iran sanctions", "iran oil",
    ],
    "Israel / Gaza / Lebanon": [
        "israel", "gaza", "hamas", "hezbollah", "ceasefire", "idf",
        "west bank", "netanyahu", "rafah", "hostage", "beirut",
        "lebanese", "occupied territory", "al-aqsa", "golan",
    ],
    "Saudi Arabia / Gulf": [
        "saudi", "mbs", "bin salman", "opec", "riyadh", "uae",
        "qatar", "abu dhabi", "bahrain", "gulf state", "aramco",
        "doha", "normaliz",
    ],
    "Broader Geopolitics": [
        "nato", "ukraine", "russia", "putin", "zelensky",
        "us-china", "taiwan", "tariff", "trump", "g7", "g20",
        "sanctions", "regime change", "coup", "military strike",
        "war", "conflict", "escalat", "de-escalat",
        "us foreign policy", "state department", "pentagon",
    ],
}

ALL_KEYWORDS = [kw for group in KEYWORD_GROUPS.values() for kw in group]

# Spreadsheet column indices (1-based, openpyxl) for Bet Pipeline tab
PIPELINE_COLS = {
    "id": 1,
    "market_question": 2,
    "spotted_by": 3,
    "date_spotted": 4,
    "pierre_prob": 5,       # left blank — filled manually
    "gabe_prob": 6,         # left blank — filled manually
    "avg_prob": 7,          # auto-calculated in sheet
    "market_price": 8,
    "implied_edge": 9,      # left blank until probs are entered
    "pierre_thesis": 10,    # left blank — filled manually
    "gabe_thesis": 11,      # left blank — filled manually
    "discussion_notes": 12,
    "break_scenarios": 13,
    "decision": 14,
    "moved_to_active": 15,
    "notes": 16,
}


# ─── API ───────────────────────────────────────────────────────────────────────

def fetch_all_markets(max_results: int = 600) -> list[dict]:
    """
    Fetch active, open markets from Polymarket Gamma API.
    Paginates automatically. Returns raw market dicts.
    """
    markets = []
    offset = 0
    batch_size = 100

    print("Fetching markets from Polymarket Gamma API...")

    while len(markets) < max_results:
        try:
            resp = requests.get(
                f"{GAMMA_API}/markets",
                params={
                    "active": "true",
                    "closed": "false",
                    "limit": batch_size,
                    "offset": offset,
                    "order": "volume",
                    "ascending": "false",
                },
                timeout=30,
            )
            resp.raise_for_status()
            batch = resp.json()

            if not batch:
                break

            markets.extend(batch)
            print(f"  ... {len(markets)} markets fetched", end="\r")

            if len(batch) < batch_size:
                break  # Last page

            offset += batch_size
            time.sleep(0.2)  # Be polite to the API

        except requests.exceptions.HTTPError as e:
            print(f"\nHTTP error at offset {offset}: {e}")
            break
        except requests.exceptions.RequestException as e:
            print(f"\nNetwork error at offset {offset}: {e}")
            break

    print(f"\nFetched {len(markets)} total markets.")
    return markets


def fetch_events_for_category(tag_slug: str, limit: int = 50) -> list[dict]:
    """
    Optionally fetch markets filtered by a Polymarket tag slug.
    Useful tag slugs: 'politics', 'world', 'middle-east', 'international-affairs'
    """
    try:
        resp = requests.get(
            f"{GAMMA_API}/events",
            params={"tag_slug": tag_slug, "limit": limit, "active": "true"},
            timeout=20,
        )
        resp.raise_for_status()
        events = resp.json()
        # Each event contains a 'markets' list — flatten
        markets = []
        for event in events:
            for m in event.get("markets", []):
                markets.append(m)
        return markets
    except requests.exceptions.RequestException:
        return []


# ─── PARSING ───────────────────────────────────────────────────────────────────

def get_yes_price(market: dict) -> float | None:
    """
    Extract YES outcome price from outcomePrices.
    outcomePrices is a JSON string like '["0.25", "0.75"]'
    where index 0 = YES price.
    """
    try:
        raw = market.get("outcomePrices", "[]")
        prices = json.loads(raw) if isinstance(raw, str) else raw
        if prices:
            price = float(prices[0])
            return price if 0.0 < price < 1.0 else None
    except (json.JSONDecodeError, ValueError, TypeError, IndexError):
        pass
    return None


def get_days_to_resolution(market: dict) -> int | None:
    """Days remaining until market resolution. Returns None if past or unknown."""
    end_str = market.get("endDate") or market.get("endDateIso")
    if not end_str:
        return None
    try:
        # Handle both 'Z' suffix and '+00:00'
        end_str = end_str.replace("Z", "+00:00")
        end_dt = datetime.fromisoformat(end_str)
        delta = (end_dt - datetime.now(timezone.utc)).days
        return delta if delta >= 0 else None
    except (ValueError, TypeError):
        return None


def get_matched_categories(market: dict) -> list[str]:
    """Return all KEYWORD_GROUPS categories this market matches."""
    text = " ".join([
        market.get("question", ""),
        market.get("description", ""),
        market.get("title", ""),
    ]).lower()

    matched = []
    for category, keywords in KEYWORD_GROUPS.items():
        if any(kw in text for kw in keywords):
            matched.append(category)
    return matched


def build_url(market: dict) -> str:
    """
    Construct the most reliable Polymarket URL for this market.
    Priority: groupSlug (parent event) → slug → conditionId → search fallback.
    """
    import urllib.parse

    # groupSlug is the parent event slug — most reliable for grouped markets
    group_slug = market.get("groupSlug") or market.get("group_slug")
    slug = market.get("slug")
    condition_id = market.get("conditionId") or market.get("condition_id")
    question = market.get("question", "")

    if group_slug:
        return f"https://polymarket.com/event/{group_slug}"
    if slug:
        return f"https://polymarket.com/event/{slug}"
    if condition_id:
        return f"https://polymarket.com/market/{condition_id}"

    # Last resort: search page using the market question
    q = urllib.parse.quote_plus(question[:100])
    return f"https://polymarket.com/search?q={q}"


# ─── FILTERING & SCORING ───────────────────────────────────────────────────────

def score_market(market: dict, yes_price: float, days: int, liquidity: float) -> int:
    """
    Compute an interest score. Higher = more worth investigating.
    This is NOT an edge estimate — it's just a prioritisation signal.
    """
    score = 0

    # Resolution timing — prefer 90-day window
    if days <= 30:
        score += 4   # Very soon; time-sensitive
    elif days <= 60:
        score += 3
    elif days <= 90:
        score += 2
    elif days <= 120:
        score += 1

    # Price distance from 50% — markets far from 50/50 are more likely mispriced
    # (either too confident or underpricing a real possibility)
    midpoint_gap = abs(yes_price - 0.5)
    if midpoint_gap >= 0.35:
        score += 3   # e.g. 85% or 15% — high potential for mispricing
    elif midpoint_gap >= 0.20:
        score += 2
    elif midpoint_gap >= 0.10:
        score += 1

    # Liquidity — more liquid = easier entry/exit
    if liquidity >= 20_000:
        score += 3
    elif liquidity >= 5_000:
        score += 2
    elif liquidity >= 1_500:
        score += 1

    # Category hits — more overlap = more central to your edge
    categories = get_matched_categories(market)
    score += len(categories)

    return score


def filter_and_score(
    markets: list[dict],
    max_days: int,
    min_liquidity: int,
    min_volume: int,
) -> list[dict]:
    """Apply filters and return scored, sorted candidate list."""
    candidates = []

    for m in markets:
        yes_price = get_yes_price(m)
        if yes_price is None:
            continue

        # Exclude near-certain markets — very little room to find edge
        if yes_price >= 0.96 or yes_price <= 0.04:
            continue

        days = get_days_to_resolution(m)
        if days is None or days > max(max_days, 180):
            continue  # Drop markets beyond extended window

        liquidity = float(m.get("liquidity", 0) or 0)
        volume = float(m.get("volume", 0) or 0)

        if liquidity < min_liquidity:
            continue
        if volume < min_volume:
            continue

        categories = get_matched_categories(m)
        if not categories:
            continue  # Not in our edge areas

        score = score_market(m, yes_price, days, liquidity)

        candidates.append({
            "id": m.get("id", ""),
            "question": m.get("question", "").strip(),
            "yes_price": yes_price,
            "market_prob_pct": round(yes_price * 100, 1),
            "days_to_resolution": days,
            "resolution_date": (m.get("endDate") or "")[:10],
            "liquidity": int(liquidity),
            "volume": int(volume),
            "categories": categories,
            "primary_category": categories[0],
            "score": score,
            "url": build_url(m),
            "market_id": m.get("id", ""),
        })

    # Sort: score descending, then days ascending (soonest first)
    candidates.sort(key=lambda x: (-x["score"], x["days_to_resolution"]))
    return candidates


# ─── HTML REPORT ───────────────────────────────────────────────────────────────

CATEGORY_COLORS = {
    "Iran": ("#c0392b", "#fdf2f2"),
    "Israel / Gaza / Lebanon": ("#2471a3", "#f0f6fb"),
    "Saudi Arabia / Gulf": ("#1e8449", "#f0faf3"),
    "Broader Geopolitics": ("#7d3c98", "#f8f4fc"),
}


def generate_html_report(candidates: list[dict], run_date: str, max_days: int) -> str:
    """Generate a clean, readable HTML digest for browser review."""

    n_within_90 = sum(1 for c in candidates if c["days_to_resolution"] <= 90)
    by_category: dict[str, list] = {}
    for c in candidates:
        by_category.setdefault(c["primary_category"], []).append(c)

    # Build table rows grouped by category
    rows_html = ""
    for cat, items in by_category.items():
        accent, bg = CATEGORY_COLORS.get(cat, ("#555", "#fafafa"))
        extra_count = len(items)
        rows_html += f"""
        <tr>
          <td colspan="7" style="background:{bg}; padding:9px 14px; font-weight:600;
              color:{accent}; font-size:12.5px; border-left:4px solid {accent};
              letter-spacing:0.3px;">
            {cat.upper()} &nbsp;·&nbsp; {extra_count} market{'s' if extra_count != 1 else ''}
          </td>
        </tr>"""

        for m in items:
            days = m["days_to_resolution"]
            if days <= 30:
                day_color = "#c0392b"
            elif days <= 60:
                day_color = "#d35400"
            elif days <= 90:
                day_color = "#27ae60"
            else:
                day_color = "#7f8c8d"

            extra_cats = ""
            if len(m["categories"]) > 1:
                extra_cats = f"""<div style="font-size:11px; color:#999; margin-top:3px;">
                    Also tagged: {', '.join(m['categories'][1:])}</div>"""

            rows_html += f"""
        <tr class="data-row">
          <td style="padding:10px 14px; font-size:13px; max-width:380px; line-height:1.45;">
            <a href="{m['url']}" target="_blank"
               style="color:#1a252f; text-decoration:none; font-weight:500;">{m['question']}</a>
            {extra_cats}
          </td>
          <td style="padding:10px 14px; text-align:center; font-weight:700; font-size:15px;
                     color:#1a252f;">{m['market_prob_pct']}%</td>
          <td style="padding:10px 14px; text-align:center;">
            <span style="background:{day_color}18; color:{day_color};
                padding:3px 9px; border-radius:10px; font-size:12px; font-weight:600;">
              {days}d
            </span>
          </td>
          <td style="padding:10px 14px; text-align:right; color:#555;
                     font-size:13px;">${m['liquidity']:,}</td>
          <td style="padding:10px 14px; text-align:right; color:#555;
                     font-size:13px;">${m['volume']:,}</td>
          <td style="padding:10px 14px; text-align:center; font-size:16px;">
            {'⭐' * min(m['score'] // 2, 5)}
          </td>
          <td style="padding:10px 14px; text-align:center;">
            <a href="{m['url']}" target="_blank"
               style="background:#2471a3; color:white; padding:4px 11px;
                      border-radius:4px; font-size:12px; text-decoration:none;
                      white-space:nowrap;">Open ↗</a>
          </td>
        </tr>"""

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <title>Polymarket Scan — {run_date}</title>
  <style>
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Arial, sans-serif;
           background: #f0f2f5; padding: 24px; color: #1a252f; }}
    .wrap {{ max-width: 1080px; margin: 0 auto; }}
    .hdr {{ background: #1a252f; color: white; padding: 22px 28px;
            border-radius: 8px 8px 0 0; }}
    .hdr h1 {{ font-size: 20px; font-weight: 700; margin-bottom: 4px; }}
    .hdr p {{ font-size: 13px; opacity: 0.6; }}
    .stats {{ display: flex; gap: 0; background: white; border-bottom: 1px solid #eaeaea; }}
    .stat {{ padding: 14px 28px; border-right: 1px solid #eaeaea; }}
    .stat .n {{ font-size: 26px; font-weight: 700; color: #1a252f; }}
    .stat .l {{ font-size: 11px; color: #888; text-transform: uppercase;
                letter-spacing: 0.6px; margin-top: 2px; }}
    .warning {{ background: #fef9e7; border-left: 4px solid #f0b429;
                padding: 12px 16px; font-size: 13px; color: #7d5a00;
                margin: 0; line-height: 1.5; }}
    table {{ width: 100%; border-collapse: collapse; background: white; }}
    thead th {{ background: #f8f9fa; padding: 9px 14px; font-size: 11px;
                text-transform: uppercase; letter-spacing: 0.5px; color: #666;
                border-bottom: 2px solid #e8e8e8; font-weight: 600; }}
    .data-row:hover td {{ background: #f7faff !important; }}
    td {{ border-bottom: 1px solid #f0f0f0; vertical-align: middle; }}
    .footer {{ text-align: center; padding: 14px; font-size: 11.5px; color: #aaa;
               background: white; border-radius: 0 0 8px 8px;
               border-top: 1px solid #eee; }}
  </style>
</head>
<body>
<div class="wrap">
  <div class="hdr">
    <h1>Polymarket Geopolitical Scan</h1>
    <p>Pierre &amp; Gabe &nbsp;·&nbsp; {run_date} &nbsp;·&nbsp; Bet Pipeline Candidates</p>
    <a href="dashboard.html" style="display:inline-block; margin-top:10px; background:rgba(255,255,255,0.15);
       color:white; padding:5px 14px; border-radius:5px; font-size:12px; text-decoration:none;">
      📊 View Dashboard →
    </a>
  </div>

  <div class="stats">
    <div class="stat"><div class="n">{len(candidates)}</div><div class="l">Candidates</div></div>
    <div class="stat"><div class="n">{n_within_90}</div><div class="l">Within 90 Days</div></div>
    <div class="stat"><div class="n">{len(by_category)}</div><div class="l">Edge Categories</div></div>
    <div class="stat"><div class="n">{max_days}d</div><div class="l">Scan Window</div></div>
  </div>

  <div class="warning">
    <strong>Process reminder:</strong> Before examining these market prices closely, sketch your
    independent probability estimate for any market you want to investigate.
    Log it in the Bet Pipeline tab (your estimate + 1-paragraph thesis) <em>before</em>
    discussing with Pierre. The numbers on this page are the market's view — not yours yet.
  </div>

  <table>
    <thead>
      <tr>
        <th style="min-width:300px; text-align:left;">Market Question</th>
        <th style="text-align:center;">Market Prob</th>
        <th style="text-align:center;">Days Left</th>
        <th style="text-align:right;">Liquidity</th>
        <th style="text-align:right;">Volume</th>
        <th style="text-align:center;">Interest</th>
        <th style="text-align:center;">Link</th>
      </tr>
    </thead>
    <tbody>
      {rows_html}
    </tbody>
  </table>

  <div class="footer">
    Generated by polymarket_scanner.py &nbsp;·&nbsp; {run_date}<br>
    Market prices shown are the <em>market's</em> implied probability — form your own estimate independently.
  </div>
</div>
</body>
</html>"""


# ─── SPREADSHEET UPDATE ────────────────────────────────────────────────────────

def update_spreadsheet(tracker_path: str, candidates: list[dict]) -> bool:
    """
    Write pipeline candidates into the Bet Pipeline tab.
    Only fills: ID, Market Question, Spotted By, Date Spotted,
                Market Price at Review, Notes (with URL + metadata).
    All thesis/probability columns are left blank for manual entry.
    """
    if not HAS_OPENPYXL:
        print("  Skipping spreadsheet update — openpyxl not installed.")
        return False

    if not os.path.exists(tracker_path):
        print(f"  Tracker not found at '{tracker_path}' — skipping spreadsheet update.")
        print("  Use --tracker path/to/your/file.xlsx")
        return False

    wb = load_workbook(tracker_path)

    if "Bet Pipeline" not in wb.sheetnames:
        print("  'Bet Pipeline' sheet not found in tracker — skipping.")
        return False

    ws = wb["Bet Pipeline"]
    today = datetime.now().strftime("%Y-%m-%d")

    # Find the first truly empty row (both col 1 and col 2 empty)
    first_empty_row = 2
    for row_idx in range(2, ws.max_row + 2):
        v1 = ws.cell(row=row_idx, column=1).value
        v2 = ws.cell(row=row_idx, column=2).value
        if v1 is None and v2 is None:
            first_empty_row = row_idx
            break

    # Determine next ID
    existing_ids = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=1, values_only=True):
        if row[0] and isinstance(row[0], (int, float)):
            existing_ids.append(int(row[0]))
    next_id = max(existing_ids) + 1 if existing_ids else 1

    # Style helpers
    normal = Font(name="Arial", size=10)
    gray = Font(name="Arial", size=10, color="999999")
    wrap = Alignment(wrap_text=True, vertical="top")

    for i, m in enumerate(candidates):
        r = first_empty_row + i

        ws.cell(r, PIPELINE_COLS["id"]).value = next_id + i
        ws.cell(r, PIPELINE_COLS["market_question"]).value = m["question"]
        ws.cell(r, PIPELINE_COLS["spotted_by"]).value = "Scanner"
        ws.cell(r, PIPELINE_COLS["date_spotted"]).value = today
        ws.cell(r, PIPELINE_COLS["market_price"]).value = m["yes_price"]
        ws.cell(r, PIPELINE_COLS["notes"]).value = (
            f"{', '.join(m['categories'])} | "
            f"{m['days_to_resolution']}d to {m['resolution_date']} | "
            f"Liq ${m['liquidity']:,} | Vol ${m['volume']:,} | "
            f"ID:{m['market_id']} | {m['url']}"
        )

        for col in range(1, 17):
            cell = ws.cell(r, col)
            cell.font = normal
            cell.alignment = wrap

        # Market price cell: format as percentage
        ws.cell(r, PIPELINE_COLS["market_price"]).number_format = "0.0%"

    wb.save(tracker_path)
    print(f"  ✓ Added {len(candidates)} rows to Bet Pipeline tab in {tracker_path}")
    return True


# ─── DELTA TRACKING & CLUSTERING ──────────────────────────────────────────────

def load_previous_data(path: str) -> list[dict]:
    """Load markets from a previous scan's markets.json for delta comparison."""
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
            return data.get("markets", [])
    except (FileNotFoundError, json.JSONDecodeError):
        return []


def compute_deltas(candidates: list[dict], previous: list[dict]) -> list[dict]:
    """
    Compare each candidate against the previous scan.
    Adds: previous_price, price_delta (signed, 0-1 scale), is_new.
    """
    prev_map = {m["question"].lower().strip(): m for m in previous}
    for c in candidates:
        key = c["question"].lower().strip()
        if key in prev_map:
            prev_price = prev_map[key].get("yes_price")
            c["previous_price"] = prev_price
            c["price_delta"] = (
                round(c["yes_price"] - prev_price, 4)
                if prev_price is not None else None
            )
            c["is_new"] = False
        else:
            c["previous_price"] = None
            c["price_delta"] = None
            c["is_new"] = True
    return candidates


def compute_clusters(candidates: list[dict]) -> list[dict]:
    """
    Assign a cluster_id to each candidate based on keyword co-occurrence.
    Markets sharing 2+ specific keywords are considered correlated.
    """
    def market_keywords(m: dict) -> set:
        text = m["question"].lower()
        hits = set()
        for kws in KEYWORD_GROUPS.values():
            for kw in kws:
                if kw in text:
                    hits.add(kw)
        return hits

    kw_sets = [(m, market_keywords(m)) for m in candidates]
    cluster_map: dict[str, int] = {}
    next_id = 0

    for i, (m, kws) in enumerate(kw_sets):
        q = m["question"]
        if q in cluster_map:
            continue
        cluster_id = next_id
        next_id += 1
        cluster_map[q] = cluster_id
        for j, (other, other_kws) in enumerate(kw_sets):
            if i == j or other["question"] in cluster_map:
                continue
            if len(kws & other_kws) >= 2:
                cluster_map[other["question"]] = cluster_id

    for c in candidates:
        c["cluster_id"] = cluster_map.get(c["question"], -1)
    return candidates


def flag_mispricing_signals(candidates: list[dict]) -> list[dict]:
    """
    Add objective mispricing signals that don't require our probability estimates.
    Signals:
      - 'sharp_move': price moved >10pp since last scan on thin liquidity
      - 'low_liquidity_outlier': price far from 50% but very low liquidity
      - 'imminent_extreme': resolves in <21 days and price is >80% or <20%
    """
    for c in candidates:
        signals = []
        delta = c.get("price_delta")
        liq = c.get("liquidity", 0)
        days = c.get("days_to_resolution", 999)
        price = c.get("yes_price", 0.5)
        mispricing = abs(price - 0.5)

        if delta is not None and abs(delta) >= 0.10 and liq < 5000:
            signals.append("sharp_move")

        if mispricing >= 0.30 and liq < 1500:
            signals.append("low_liquidity_outlier")

        if days <= 21 and (price >= 0.80 or price <= 0.20):
            signals.append("imminent_extreme")

        c["mispricing_signals"] = signals
    return candidates


def save_markets_json(candidates: list[dict], output_dir: str, run_date: str) -> str:
    """Save full market data as markets.json for the dashboard to consume."""
    data = {
        "generated_at": run_date,
        "total": len(candidates),
        "markets": candidates,
    }
    path = os.path.join(output_dir, "markets.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, default=str)
    return path


# ─── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Polymarket geopolitical market scanner — Pierre & Gabe pipeline"
    )
    parser.add_argument(
        "--tracker", default="polymarket_tracker_v2.xlsx",
        help="Path to tracker xlsx (default: polymarket_tracker_v2.xlsx)"
    )
    parser.add_argument(
        "--output-dir", default=".",
        help="Directory to save the HTML report (default: current directory)"
    )
    parser.add_argument(
        "--days", type=int, default=90,
        help="Max days to resolution (default: 90; use 180 for extended scan)"
    )
    parser.add_argument(
        "--min-liquidity", type=int, default=500,
        help="Minimum market liquidity in USD (default: 500)"
    )
    parser.add_argument(
        "--min-volume", type=int, default=1000,
        help="Minimum traded volume in USD (default: 1000)"
    )
    parser.add_argument(
        "--top", type=int, default=40,
        help="Max candidates to include (default: 40)"
    )
    parser.add_argument(
        "--no-spreadsheet", action="store_true",
        help="Produce report only; do not write to spreadsheet"
    )
    parser.add_argument(
        "--previous-data", default=None,
        help="Path to previous scan's markets.json for delta tracking"
    )
    args = parser.parse_args()

    run_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    print(f"\n{'='*55}")
    print(f"  Polymarket Scanner | {run_date}")
    print(f"  Window: {args.days}d | Min liq: ${args.min_liquidity:,} | Top: {args.top}")
    print(f"{'='*55}\n")

    # Step 1: Fetch
    markets = fetch_all_markets(max_results=600)
    if not markets:
        print("No markets fetched. Check your internet connection.")
        sys.exit(1)

    # Step 2: Filter and score
    print("Filtering for geopolitical relevance...")
    candidates = filter_and_score(
        markets,
        max_days=args.days,
        min_liquidity=args.min_liquidity,
        min_volume=args.min_volume,
    )
    candidates = candidates[: args.top]

    if not candidates:
        print(
            "\nNo candidates found with current filters.\n"
            "Try: --days 180 or --min-liquidity 200 or --min-volume 500"
        )
        sys.exit(0)

    # Step 3: Delta tracking, clustering, mispricing signals
    previous = load_previous_data(args.previous_data) if args.previous_data else []
    if previous:
        print(f"  Loaded {len(previous)} markets from previous scan for delta tracking.")
    candidates = compute_deltas(candidates, previous)
    candidates = compute_clusters(candidates)
    candidates = flag_mispricing_signals(candidates)

    print(f"\nFound {len(candidates)} pipeline candidates:\n")
    for c in candidates[:10]:
        sig = " ⚑" if c.get("mispricing_signals") else ""
        print(f"  [{c['primary_category'][:25]:<25}] "
              f"{c['market_prob_pct']:>5.1f}% | {c['days_to_resolution']:>3}d | "
              f"{c['question'][:55]}{sig}")
    if len(candidates) > 10:
        print(f"  ... and {len(candidates) - 10} more (see HTML report)")

    # Step 4: Save markets.json for dashboard
    Path(args.output_dir).mkdir(parents=True, exist_ok=True)
    json_path = save_markets_json(candidates, args.output_dir, run_date)
    print(f"\n✓ Market data saved: {json_path}")

    # Step 5: Generate HTML report
    report_name = "index.html"
    report_path = os.path.join(args.output_dir, report_name)
    html = generate_html_report(candidates, run_date, args.days)
    Path(report_path).write_text(html, encoding="utf-8")
    print(f"✓ Scan report saved: {report_path}")

    # Step 6: Update spreadsheet
    if not args.no_spreadsheet:
        print(f"Updating spreadsheet: {args.tracker}")
        update_spreadsheet(args.tracker, candidates)

    print(f"\n{'='*55}")
    print("  Done. Next steps:")
    print("  1. Open index.html in your browser (scan report)")
    print("  2. Open dashboard.html for visualisations")
    print("  3. For any market you want to investigate:")
    print("     → Form YOUR probability estimate BEFORE looking at the price")
    print("     → Write a 1-paragraph thesis in the Bet Pipeline tab")
    print("     → Then compare with Pierre at your next session")
    print(f"{'='*55}\n")


if __name__ == "__main__":
    main()
